from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


async def parse_schedule(date: int):
    lessons = []
    wb = load_workbook('schedule.xlsx')
    ws = wb.active
    num = 1
    for col in range(2, 5):
        char = get_column_letter(col)
        lesson = (ws[char + str(date)]).value
        if lesson is not None:
            lessons.append(
                f'{num}. {lesson}\n'
            )
            num += 1
    return lessons
